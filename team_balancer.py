#!/usr/bin/env python3
"""
FC 와디즈 팀 밸런서
==================
축구 선수들의 능력치를 기반으로 균형잡힌 팀을 자동으로 구성합니다.
"""

import sys
from pathlib import Path
from itertools import combinations
from typing import Dict, List, Tuple, Optional
import math

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl 라이브러리가 필요합니다!")
    print("   설치 방법: pip install openpyxl")
    sys.exit(1)


# ANSI 색상 코드
class Color:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    BLACK_TEAM = "\033[90m"  # 진한 회색
    WHITE_TEAM = "\033[97m"  # 밝은 흰색
    HEADER = "\033[95m"
    BLUE = "\033[94m"
    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"


# 능력치 카테고리 정의
STAT_CATEGORIES = {
    "공격력": ["슛파워", "슛정확도", "골 결정력"],
    "미드필드": ["패스", "위치선정"],
    "수비력": ["수비", "몸싸움"],
    "피지컬": ["체력", "속력", "가속도"],
    "멘탈": ["정신력", "적극성"]
}


class Player:
    """선수 클래스"""

    def __init__(self, name: str, player_type: str, stats: Dict[str, float]):
        self.name = name
        self.player_type = player_type
        self.stats = stats
        self.total = sum(stats.values())

    def get_category_total(self, category: str) -> float:
        """특정 카테고리의 능력치 합계"""
        stat_names = STAT_CATEGORIES.get(category, [])
        return sum(self.stats.get(stat, 0) for stat in stat_names)

    def __repr__(self):
        return f"{self.name} ({self.player_type}) - {self.total:.0f}"


class TeamBalancer:
    """팀 밸런서 클래스"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.players: List[Player] = []
        self.stat_names: List[str] = []

    def load_data(self) -> bool:
        """엑셀 파일에서 선수 데이터 로드"""
        try:
            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb.active

            # 선수 이름 읽기 (B2:J2) - row 1은 비어있고 이름은 row 2
            player_names = []
            for col in range(2, 11):  # B(2) to J(10)
                name = ws.cell(row=2, column=col).value
                if name:
                    player_names.append(name.strip())

            # 능력치 이름 읽기 (A3:A14, 쇼맨쉽 제외)
            stat_names = []
            for row in range(3, 15):  # 3-14
                stat_name = ws.cell(row=row, column=1).value
                if stat_name and stat_name.strip() != "쇼맨쉽":
                    stat_names.append(stat_name.strip())

            self.stat_names = stat_names

            # 선수 유형 읽기 (B18:J18)
            player_types = []
            for col in range(2, 11):
                player_type = ws.cell(row=18, column=col).value
                if player_type:
                    player_types.append(player_type.strip())
                else:
                    player_types.append("일반")

            # 선수별 능력치 파싱
            for idx, name in enumerate(player_names):
                col = idx + 2  # B=2, C=3, ...
                stats = {}

                for stat_idx, stat_name in enumerate(self.stat_names):
                    row = stat_idx + 3  # 3부터 시작 (row 1 빈행, row 2 이름)
                    value = ws.cell(row=row, column=col).value
                    try:
                        stats[stat_name] = float(value) if value else 0.0
                    except (ValueError, TypeError):
                        stats[stat_name] = 0.0

                player = Player(
                    name=name,
                    player_type=player_types[idx] if idx < len(player_types) else "일반",
                    stats=stats
                )
                self.players.append(player)

            wb.close()
            return True

        except FileNotFoundError:
            print(f"{Color.RED}❌ 파일을 찾을 수 없습니다: {self.excel_path}{Color.RESET}")
            return False
        except Exception as e:
            print(f"{Color.RED}❌ 데이터 로드 중 오류 발생: {e}{Color.RESET}")
            return False

    def display_all_players(self):
        """모든 선수 목록 표시"""
        print(f"\n{Color.HEADER}{'='*80}{Color.RESET}")
        print(f"{Color.BOLD}📋 FC 와디즈 선수 명단{Color.RESET}")
        print(f"{Color.HEADER}{'='*80}{Color.RESET}\n")

        # 테이블 헤더
        print(f"{Color.CYAN}{'No':<4} {'이름':<8} {'유형':<15} {'총합':<6} {'공격':<6} {'미드':<6} {'수비':<6} {'체력':<6}{Color.RESET}")
        print(f"{'-'*80}")

        # 선수 정보
        for idx, player in enumerate(self.players, 1):
            attack = player.get_category_total("공격력")
            mid = player.get_category_total("미드필드")
            defense = player.get_category_total("수비력")
            physical = player.get_category_total("피지컬")

            print(f"{idx:<4} {player.name:<8} {player.player_type:<15} "
                  f"{player.total:<6.0f} {attack:<6.0f} {mid:<6.0f} {defense:<6.0f} {physical:<6.0f}")

        print(f"{'-'*80}\n")

    def select_players(self) -> List[Player]:
        """참여할 선수 선택"""
        print(f"{Color.BOLD}👥 오늘 참여할 선수를 선택하세요{Color.RESET}")
        print(f"   (번호를 입력하고 Enter, 완료하면 'done' 입력)\n")

        selected_indices = set()

        while True:
            try:
                user_input = input(f"{Color.GREEN}선수 번호 (또는 'done'): {Color.RESET}").strip()

                if user_input.lower() == 'done':
                    break

                # 범위 입력 지원 (예: 1-5)
                if '-' in user_input:
                    start, end = map(int, user_input.split('-'))
                    for i in range(start, end + 1):
                        if 1 <= i <= len(self.players):
                            selected_indices.add(i)
                    print(f"   ✅ {start}번부터 {end}번까지 선택됨")
                else:
                    idx = int(user_input)
                    if 1 <= idx <= len(self.players):
                        if idx in selected_indices:
                            selected_indices.remove(idx)
                            print(f"   ❌ {self.players[idx-1].name} 선택 해제")
                        else:
                            selected_indices.add(idx)
                            print(f"   ✅ {self.players[idx-1].name} 선택됨")
                    else:
                        print(f"   {Color.RED}⚠️  1부터 {len(self.players)} 사이의 숫자를 입력하세요{Color.RESET}")

                # 현재 선택된 선수 표시
                if selected_indices:
                    selected_names = [self.players[i-1].name for i in sorted(selected_indices)]
                    print(f"   {Color.CYAN}현재 선택: {', '.join(selected_names)} ({len(selected_names)}명){Color.RESET}\n")

            except ValueError:
                if user_input.lower() != 'done':
                    print(f"   {Color.RED}⚠️  올바른 번호를 입력하세요{Color.RESET}")
            except KeyboardInterrupt:
                print(f"\n{Color.YELLOW}취소되었습니다.{Color.RESET}")
                sys.exit(0)

        if not selected_indices:
            print(f"{Color.YELLOW}⚠️  선수가 선택되지 않았습니다. 전체 선수로 진행합니다.{Color.RESET}")
            return self.players

        return [self.players[i-1] for i in sorted(selected_indices)]

    def select_team_size(self, available_count: int) -> int:
        """팀 사이즈 선택"""
        max_team_size = available_count // 2

        print(f"\n{Color.BOLD}⚽ 팀 구성을 선택하세요{Color.RESET}")
        print(f"   (참여 선수: {available_count}명, 최대 {max_team_size}:{max_team_size} 가능)\n")

        valid_sizes = [i for i in [3, 4, 5, 6] if i <= max_team_size]

        for size in valid_sizes:
            print(f"   {size}. {size}:{size} ({size*2}명)")

        while True:
            try:
                choice = input(f"\n{Color.GREEN}선택: {Color.RESET}").strip()
                size = int(choice)

                if size in valid_sizes:
                    return size
                else:
                    print(f"{Color.RED}⚠️  {', '.join(map(str, valid_sizes))} 중 하나를 선택하세요{Color.RESET}")
            except ValueError:
                print(f"{Color.RED}⚠️  올바른 숫자를 입력하세요{Color.RESET}")
            except KeyboardInterrupt:
                print(f"\n{Color.YELLOW}취소되었습니다.{Color.RESET}")
                sys.exit(0)

    def lock_players(self, players: List[Player]) -> Tuple[List[Player], List[Player]]:
        """특정 선수를 특정 팀에 고정"""
        print(f"\n{Color.BOLD}🔒 특정 선수를 팀에 고정하시겠습니까? (선택사항){Color.RESET}")
        print(f"   (y/n): ", end="")

        response = input().strip().lower()

        if response != 'y':
            return [], []

        black_locked = []
        white_locked = []

        print(f"\n{Color.BLACK_TEAM}⚫ 블랙팀에 고정할 선수{Color.RESET} (번호 입력, 없으면 Enter):")
        while True:
            try:
                user_input = input(f"   선수 번호 (또는 Enter): ").strip()
                if not user_input:
                    break

                idx = int(user_input)
                if 1 <= idx <= len(players):
                    player = players[idx-1]
                    if player not in black_locked and player not in white_locked:
                        black_locked.append(player)
                        print(f"   ✅ {player.name} → ⚫ 블랙팀 고정")
                else:
                    print(f"   {Color.RED}⚠️  1부터 {len(players)} 사이의 숫자를 입력하세요{Color.RESET}")
            except ValueError:
                break

        print(f"\n{Color.WHITE_TEAM}⚪ 화이트팀에 고정할 선수{Color.RESET} (번호 입력, 없으면 Enter):")
        while True:
            try:
                user_input = input(f"   선수 번호 (또는 Enter): ").strip()
                if not user_input:
                    break

                idx = int(user_input)
                if 1 <= idx <= len(players):
                    player = players[idx-1]
                    if player not in black_locked and player not in white_locked:
                        white_locked.append(player)
                        print(f"   ✅ {player.name} → ⚪ 화이트팀 고정")
                else:
                    print(f"   {Color.RED}⚠️  1부터 {len(players)} 사이의 숫자를 입력하세요{Color.RESET}")
            except ValueError:
                break

        return black_locked, white_locked

    def calculate_balance_score(self, team1: List[Player], team2: List[Player]) -> Tuple[float, Dict]:
        """팀 밸런스 점수 계산 (낮을수록 균형잡힘)"""
        # 총합 차이
        total1 = sum(p.total for p in team1)
        total2 = sum(p.total for p in team2)
        total_diff = abs(total1 - total2)

        # 카테고리별 차이
        category_diffs = {}
        total_category_diff = 0

        for category in STAT_CATEGORIES.keys():
            cat1 = sum(p.get_category_total(category) for p in team1)
            cat2 = sum(p.get_category_total(category) for p in team2)
            diff = abs(cat1 - cat2)
            category_diffs[category] = diff
            total_category_diff += diff

        # 최종 점수 = 총합 차이 + 카테고리 차이의 가중 평균
        score = total_diff + (total_category_diff * 0.5)

        details = {
            "total_diff": total_diff,
            "category_diffs": category_diffs,
            "team1_total": total1,
            "team2_total": total2
        }

        return score, details

    def find_balanced_teams(self, players: List[Player], team_size: int,
                           black_locked: List[Player] = None,
                           white_locked: List[Player] = None) -> List[Tuple]:
        """균형잡힌 팀 조합 찾기"""
        black_locked = black_locked or []
        white_locked = white_locked or []

        # 고정되지 않은 선수들
        flexible_players = [p for p in players if p not in black_locked and p not in white_locked]

        # 필요한 추가 선수 수
        black_needed = team_size - len(black_locked)
        white_needed = team_size - len(white_locked)

        if black_needed < 0 or white_needed < 0:
            print(f"{Color.RED}❌ 고정된 선수가 팀 사이즈보다 많습니다!{Color.RESET}")
            return []

        if black_needed + white_needed > len(flexible_players):
            print(f"{Color.RED}❌ 선수 수가 부족합니다!{Color.RESET}")
            return []

        print(f"\n{Color.CYAN}🔍 최적의 팀 조합을 찾는 중...{Color.RESET}", end="", flush=True)

        results = []

        # 가능한 모든 조합 시도
        for black_combo in combinations(flexible_players, black_needed):
            black_team = list(black_locked) + list(black_combo)
            remaining = [p for p in flexible_players if p not in black_combo]

            # 화이트팀은 남은 선수에서 선택
            if len(remaining) >= white_needed:
                for white_combo in combinations(remaining, white_needed):
                    white_team = list(white_locked) + list(white_combo)

                    score, details = self.calculate_balance_score(black_team, white_team)
                    results.append((score, black_team, white_team, details))

        print(f" {len(results)}개 조합 분석 완료! ✅\n")

        # 점수 기준 정렬
        results.sort(key=lambda x: x[0])

        return results[:3]  # 상위 3개 반환

    def display_team_comparison(self, rank: int, black_team: List[Player],
                               white_team: List[Player], details: Dict):
        """팀 비교 표시"""
        print(f"\n{Color.HEADER}{'='*80}{Color.RESET}")
        print(f"{Color.BOLD}🏆 추천 조합 #{rank}{Color.RESET}")
        print(f"{Color.HEADER}{'='*80}{Color.RESET}\n")

        # 팀 구성원 표시
        print(f"{Color.BLACK_TEAM}{'⚫ 블랙팀':<40}{Color.RESET}  {Color.WHITE_TEAM}{'⚪ 화이트팀':<40}{Color.RESET}")
        print(f"{'-'*80}")

        max_players = max(len(black_team), len(white_team))
        for i in range(max_players):
            black_info = ""
            white_info = ""

            if i < len(black_team):
                p = black_team[i]
                black_info = f"{p.name:<8} ({p.player_type:<12}) {p.total:>5.0f}"

            if i < len(white_team):
                p = white_team[i]
                white_info = f"{p.name:<8} ({p.player_type:<12}) {p.total:>5.0f}"

            print(f"{Color.BLACK_TEAM}{black_info:<40}{Color.RESET}  {Color.WHITE_TEAM}{white_info:<40}{Color.RESET}")

        print(f"{'-'*80}\n")

        # 카테고리별 비교
        print(f"{Color.CYAN}📊 카테고리별 능력치 비교{Color.RESET}\n")
        print(f"{'카테고리':<12} {'블랙팀':>10} {'화이트팀':>10} {'차이':>10} {'밸런스':>15}")
        print(f"{'-'*80}")

        for category in STAT_CATEGORIES.keys():
            black_cat = sum(p.get_category_total(category) for p in black_team)
            white_cat = sum(p.get_category_total(category) for p in white_team)
            diff = abs(black_cat - white_cat)

            # 밸런스 바 표시
            max_val = max(black_cat, white_cat)
            if max_val > 0:
                balance_pct = (1 - diff / max_val) * 100
                bar_length = int(balance_pct / 10)
                balance_bar = "█" * bar_length + "░" * (10 - bar_length)
            else:
                balance_bar = "░" * 10

            print(f"{category:<12} {black_cat:>10.0f} {white_cat:>10.0f} "
                  f"{diff:>10.0f} {balance_bar} {balance_pct:.0f}%")

        # 총합 비교
        print(f"{'-'*80}")
        total_diff = details["total_diff"]
        black_total = details["team1_total"]
        white_total = details["team2_total"]

        overall_balance = (1 - total_diff / max(black_total, white_total)) * 100 if max(black_total, white_total) > 0 else 100

        print(f"{Color.BOLD}{'총합':<12} {black_total:>10.0f} {white_total:>10.0f} "
              f"{total_diff:>10.0f} {'전체 밸런스:'} {overall_balance:.1f}%{Color.RESET}\n")

        # 밸런스 평가
        if overall_balance >= 95:
            grade = f"{Color.GREEN}완벽! 🌟{Color.RESET}"
        elif overall_balance >= 90:
            grade = f"{Color.GREEN}매우 좋음 👍{Color.RESET}"
        elif overall_balance >= 85:
            grade = f"{Color.CYAN}좋음 ✓{Color.RESET}"
        elif overall_balance >= 80:
            grade = f"{Color.YELLOW}보통 -{Color.RESET}"
        else:
            grade = f"{Color.RED}불균형 ⚠️{Color.RESET}"

        print(f"💯 밸런스 평가: {grade}\n")


def main():
    """메인 실행 함수"""
    print(f"\n{Color.HEADER}")
    print("╔════════════════════════════════════════════════════════════════════════════╗")
    print("║                                                                            ║")
    print("║                    ⚽  FC 와디즈 팀 밸런서  ⚽                              ║")
    print("║                                                                            ║")
    print("║              균형잡힌 축구 팀을 자동으로 구성해드립니다!                   ║")
    print("║                                                                            ║")
    print("╚════════════════════════════════════════════════════════════════════════════╝")
    print(f"{Color.RESET}\n")

    # Excel 파일 경로
    excel_path = Path(__file__).parent / "FC와디즈_능력치.xlsx"

    if not excel_path.exists():
        print(f"{Color.RED}❌ 'FC와디즈_능력치.xlsx' 파일을 찾을 수 없습니다!{Color.RESET}")
        print(f"   위치: {excel_path}")
        sys.exit(1)

    # 팀 밸런서 초기화
    balancer = TeamBalancer(str(excel_path))

    # 데이터 로드
    if not balancer.load_data():
        sys.exit(1)

    print(f"{Color.GREEN}✅ {len(balancer.players)}명의 선수 데이터 로드 완료!{Color.RESET}")

    # 전체 선수 표시
    balancer.display_all_players()

    # 참여 선수 선택
    selected_players = balancer.select_players()
    print(f"\n{Color.GREEN}✅ {len(selected_players)}명 선택됨{Color.RESET}")

    # 팀 사이즈 선택
    team_size = balancer.select_team_size(len(selected_players))
    print(f"\n{Color.GREEN}✅ {team_size}:{team_size} 구성 선택됨{Color.RESET}")

    # 선수 고정 (선택사항)
    black_locked, white_locked = balancer.lock_players(selected_players)

    # 균형잡힌 팀 찾기
    results = balancer.find_balanced_teams(selected_players, team_size, black_locked, white_locked)

    if not results:
        print(f"{Color.RED}❌ 조건에 맞는 팀 조합을 찾을 수 없습니다.{Color.RESET}")
        sys.exit(1)

    # 결과 표시
    for idx, (score, black_team, white_team, details) in enumerate(results, 1):
        balancer.display_team_comparison(idx, black_team, white_team, details)

    print(f"{Color.HEADER}{'='*80}{Color.RESET}")
    print(f"{Color.BOLD}🎉 팀 밸런싱 완료! 좋은 경기 되세요! ⚽{Color.RESET}")
    print(f"{Color.HEADER}{'='*80}{Color.RESET}\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n{Color.YELLOW}👋 프로그램을 종료합니다.{Color.RESET}\n")
        sys.exit(0)
